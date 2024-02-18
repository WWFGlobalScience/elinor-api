from copy import copy
from django.db import transaction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, DEFAULT_FONT, Protection
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.protection import SheetProtection
from zipfile import BadZipFile

from ..models import SurveyQuestionLikert, SurveyAnswerLikert
from ..resources.survey import SurveyAnswerLikertSerializer
from ..utils import strip_html
from ..utils.assessment import (
    assessment_xlsx_has_errors,
    enforce_required_attributes,
    questionlikerts,
)
from . import (
    ingest_400,
    ERROR,
    MISSING_SHEET,
    INVALID_HEADER,
    INVALID_HEADER_CELLS,
    INVALID_FILE_LOAD,
    ASSESSMENT_ID_MISMATCH,
    INVALID_SHEET,
    INVALID_QUESTIONS,
    INVALID_CHOICES,
    SURVEYANSWERLIKERTSERIALIZER,
    ANSWER_SAVE,
)


# TODO: create elinordata.org subdomain for this s3 bucket
DOCUMENTATION_URL = "https://elinor-user-files.s3.amazonaws.com/dev/Document/2/Elinor_assessment_tool_protocol_v2022.1.pdf"
bold = copy(DEFAULT_FONT)
bold.bold = True
bold14pt = copy(DEFAULT_FONT)
bold14pt.bold = True
bold14pt.size = 14
wrapped_alignment = Alignment(
    horizontal="general",
    vertical="top",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
)

WS_DEF = {
    "survey": {
        "title": {
            "row": 1,
            "header": [
                {
                    "content": "",
                    "font": bold14pt,
                },
                {},
            ],
        },
        "intro": {
            "row": 2,
            "header": [
                {
                    "content": "Please make sure you read our protocol before answering this survey:",
                    "font": bold,
                },
                {},
                {"content": DOCUMENTATION_URL, "hyperlink": DOCUMENTATION_URL},
            ],
        },
        "columns": {
            "row": 4,
            "header": [
                {"content": "Survey Question", "font": bold14pt, "width": 80},
                {"content": "key", "hidden": True},
                {"content": "Answer", "width": 40},
                {"content": "Explanation", "width": 40},
                {"content": "Rationale", "width": 20},
                {"content": "Information", "width": 20},
                {"content": "Guidance", "width": 20},
            ],
        },
    },
    "choices": {
        "columns": {
            "row": 1,
            "header": [
                {"content": "key"},
                {"content": "excellent_3"},
                {"content": "good_2"},
                {"content": "average_1"},
                {"content": "poor_0"},
            ],
        }
    },
}


class InvalidChoice(Exception):
    pass


def get_choice_by_answer(question, choice):
    for question_choice in question.choices:
        val = int(question_choice.split(": ")[0])
        if choice == val:
            return question_choice
    return ""


def get_user_choice(choice_str):
    if isinstance(choice_str, int):
        return choice_str

    if not isinstance(choice_str, str):
        return None

    choice = choice_str.split(":")[0]
    try:
        choice_val = int(choice)
        return choice_val
    except:
        raise InvalidChoice


class AssessmentXLSX:
    def __init__(self, assessment):
        self.assessment = assessment
        enforce_required_attributes(self.assessment)
        self._questions = []
        self._answers = {}
        self._ws_survey = None
        self._ws_choices = None
        self.workbook = None
        self.xlsxfile = None
        self.validations = {}
        self.ws_def = WS_DEF
        self.cols = {
            "key": self.get_survey_col("key"),
            "keyl": get_column_letter(self.get_survey_col("key") + 1),
            "answer": self.get_survey_col("answer"),
            "answerl": get_column_letter(self.get_survey_col("answer") + 1),
            "explanation": self.get_survey_col("explanation"),
            "explanationl": get_column_letter(self.get_survey_col("explanation") + 1),
        }

    @property
    def sheetnames(self):
        return list(self.ws_def.keys())

    def _set_sheet(self, sheetprop, sheetnum):
        if not self.workbook:
            raise ReferenceError("workbook must be defined before accessing worksheets")

        sheetname = self.sheetnames[sheetnum]
        try:
            sheet = self.workbook.get_sheet_by_name(sheetname)
            setattr(self, sheetprop, sheet)
        except KeyError:
            error = ingest_400(
                MISSING_SHEET,
                f"missing sheet with name '{sheetname}'",
                {"sheetname": sheetname},
            )
            self.validations.update(error)

    def get_survey_col(self, column_name):
        for i, cell in enumerate(self.ws_def[self.sheetnames[0]]["columns"]["header"]):
            if cell["content"].lower() == column_name:
                return i
        return None

    @property
    def ws_survey(self):
        if not self._ws_survey:
            self._set_sheet("_ws_survey", 0)
        return self._ws_survey

    @property
    def ws_choices(self):
        if not self._ws_choices:
            self._set_sheet("_ws_choices", 1)
        return self._ws_choices

    # Note: self.questions is based on all SurveyQuestionLikert objects in the db at the time of access,
    # regardless of whether they're in an attribute selected for the assessment. This allows the
    # user to answer questions that are not part of the assessment (though they will not be scored).
    @property
    def questions(self):
        if not self._questions:
            qs = questionlikerts()
            _choices_header = self.ws_def[self.sheetnames[1]]["columns"]["header"]
            for q in qs:
                question = copy(q)
                question.choices = []
                for col in _choices_header[1:]:
                    attr = col.get("content")
                    val = getattr(q, attr)
                    choice = attr.split("_")[1]
                    question.choices.append(f"{choice}: {strip_html(val)}")
                self._questions.append(question)

        return self._questions

    @property
    def answers(self):
        if not self._answers:
            answers = (
                SurveyAnswerLikert.objects.filter(assessment=self.assessment)
                .select_related("question", "question__attribute")
                .order_by(
                    "question__attribute__order",
                    "question__attribute__name",
                    "question__number",
                )
            )

            for a in answers:
                self._answers[a.question.key] = {
                    "choice": a.choice,
                    "explanation": a.explanation,
                }

        return self._answers

    @answers.setter
    def answers(self, answers_dict):
        self._answers = answers_dict

    def write_header(self, sheetname, section="columns"):
        _header_row = self.ws_def[sheetname][section]["row"]
        _header = self.ws_def[sheetname][section]["header"]
        ws = self.workbook.get_sheet_by_name(sheetname)

        for i, col in enumerate(_header):
            _cell = ws.cell(row=_header_row, column=i + 1, value=col.get("content"))
            col_index = get_column_letter(i + 1)
            _font = col.get("font")
            _link = col.get("hyperlink")
            _hidden = col.get("hidden")
            _width = col.get("width")
            if _font:
                _cell.font = _font
            if _link:
                _cell.hyperlink = _link
            if _hidden:
                ws.column_dimensions[col_index].hidden = True
            if _width:
                ws.column_dimensions[col_index].width = _width

    def validate_header(self, sheetname):
        sheet = self.workbook.get_sheet_by_name(sheetname)
        header_row = self.ws_def[sheetname]["columns"]["row"]
        user_header_row = list(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        if not user_header_row:
            error = ingest_400(
                INVALID_HEADER,
                f"no header found for sheet {sheetname}",
                {"sheetname": sheetname},
            )
            self.validations.update(error)
            return
        user_header = user_header_row[0]

        header_error_cells = []
        for i, user_cell in enumerate(user_header, start=1):
            col = get_column_letter(i)
            header_cell = self.ws_def[sheetname]["columns"]["header"][i - 1]
            if user_cell != header_cell["content"]:
                header_error_cells.append(f"{col}{header_row}")

        if header_error_cells:
            error = ingest_400(
                INVALID_HEADER_CELLS,
                f"invalid headers in cells: {','.join(header_error_cells)}",
                {"header_error_cells": header_error_cells},
            )
            self.validations.update(error)

    def add_survey_validation(self, qkey):
        val_formula = ""
        # Could use question.choices but doesn't work if a choice has a comma
        for i, row in enumerate(self.ws_choices.iter_rows()):
            key = row[0].value
            if key == qkey:
                val_formula = f"{self.sheetnames[1]}!$B${i + 1}:$E${i + 1}"
        validation = DataValidation(
            type="list",
            allow_blank=True,
            formula1=val_formula,
            showErrorMessage=True,
            errorTitle="invalid choice",
            error="Please select a choice from the list",
        )
        self.ws_survey.add_data_validation(validation)
        return validation

    # noinspection PyMethodMayBeStatic
    def protect_sheet(self, ws, exception_columns=None):
        # self.ws_choices.protection = SheetProtection(formatColumns=False, formatRows=False, formatCells=False)
        # TODO: setting formatColumns/formatRows = False has no effect (at least in LibreOffice).
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        for col in exception_columns or []:
            for cell in ws[col]:
                cell.protection = Protection(locked=False)

    def get_question_by_key(self, key):
        for question in self.questions:
            if question.key == key:
                return question
        return None

    def generate_from_assessment(self):
        self.workbook = Workbook(iso_dates=True)
        self.workbook.worksheets[0].title = self.sheetnames[0]
        self.workbook.create_sheet(self.sheetnames[1])

        self.write_header(self.sheetnames[0], section="title")
        titlecrow = self.ws_def[self.sheetnames[0]]["title"]["row"]
        self.ws_survey.cell(row=titlecrow, column=1, value=self.assessment.name)
        self.ws_survey.cell(row=titlecrow, column=2, value=self.assessment.pk)
        self.write_header(self.sheetnames[0], section="intro")
        self.write_header(self.sheetnames[0])
        self.write_header(self.sheetnames[1])
        for q in self.questions:
            choices = [q.key] + [c for c in q.choices]
            self.ws_choices.append(choices)

        arow = self.ws_def[self.sheetnames[0]]["columns"]["row"] + 1
        for attribute in self.assessment.attributes.order_by("order", "name"):
            attr_cell = self.ws_survey.cell(
                row=arow, column=1, value=attribute.name.upper()
            )
            attr_cell.font = bold
            attr_cell.alignment = wrapped_alignment
            arow += 1
            for question in self.questions:
                if question.attribute == attribute:
                    qtext = f"{question.number}. {question.text}"
                    answer = self.answers.get(question.key, {}) or {}
                    choice = answer.get("choice", "")
                    choice_text = get_choice_by_answer(question, choice)
                    validation = self.add_survey_validation(question.key)
                    explanation = answer.get("explanation", "")
                    rationale = strip_html(question.rationale)
                    information = strip_html(question.information)
                    guidance = strip_html(question.guidance)

                    self.ws_survey.row_dimensions[arow].height = 32
                    qrow = [
                        qtext,
                        question.key,
                        choice_text,
                        explanation,
                        rationale,
                        information,
                        guidance,
                    ]
                    for i, val in enumerate(qrow):
                        _cell = self.ws_survey.cell(row=arow, column=i + 1, value=val)
                        if i == 0:
                            _cell.alignment = wrapped_alignment
                        if i == 2:
                            validation.add(_cell)

                    arow += 1

        self.protect_sheet(self.ws_choices)
        self.protect_sheet(self.ws_survey, ["C", "D"])

    def check_file_structure(self, file):
        try:
            self.workbook = load_workbook(file, read_only=True)
        except (InvalidFileException, BadZipFile):
            # openpyxl uses zipfile to read file; if the file has an xlsx extension but isn't one, the
            # exception message is about a zip file, which would be confusing to users. So we'll return our own.
            error = ingest_400(INVALID_FILE_LOAD, "invalid xlsx file")
            self.validations.update(error)

        titlecrow = self.ws_def[self.sheetnames[0]]["title"]["row"]
        user_assessmment_id = self.ws_survey.cell(row=titlecrow, column=2).value
        if user_assessmment_id != self.assessment.pk:
            error = ingest_400(
                ASSESSMENT_ID_MISMATCH,
                f"assessment id {user_assessmment_id} in B{titlecrow} does not match requested assessment {self.assessment.pk}",
                {
                    "user_assessmment_id": user_assessmment_id,
                    "cell": f"B{titlecrow}",
                    "assessment_id": self.assessment.pk,
                },
            )
            self.validations.update(error)

        try:
            for sheet in self.ws_def.keys():
                self.validate_header(sheet)
        except KeyError as e:
            error = ingest_400(INVALID_SHEET, str(e).strip("'"))
            self.validations.update(error)

    def load_from_file(self, file):
        self.check_file_structure(file)
        if assessment_xlsx_has_errors(self):
            return

        header_row = self.ws_def[self.sheetnames[0]]["columns"]["row"]
        start_row = header_row + 1
        rows = self.ws_survey.iter_rows(min_row=start_row, values_only=True)
        user_answers = {}
        question_error_cells = []
        choice_error_cells = []
        for i, row in enumerate(rows, start=start_row):
            key = row[self.cols["key"]]
            if not key:
                continue
            question = self.get_question_by_key(key)
            if not question:
                question_error_cells.append(f"{self.cols['keyl']}{i}")
                continue

            user_answer = row[self.cols["answer"]]
            try:
                choice = get_user_choice(user_answer)
                user_answers[key] = {
                    "choice": choice,
                    "explanation": str(row[self.cols["explanation"]] or ""),
                }
            except InvalidChoice:
                choice_error_cells.append(f"{self.cols['answerl']}{i}")

        if question_error_cells:
            error = ingest_400(
                INVALID_QUESTIONS,
                f"invalid question keys in cells: {','.join(question_error_cells)}",
                {"question_error_cells": question_error_cells},
            )
            self.validations.update(error)
        if choice_error_cells:
            error = ingest_400(
                INVALID_CHOICES,
                f"invalid choices in cells: {','.join(choice_error_cells)}",
                {"choice_error_cells": choice_error_cells},
            )
            self.validations.update(error)

        self.answers = user_answers

    def submit_answers(self, dryrun):
        answer_serializers = []
        answer_serializer_errors = []
        for key in self.answers.keys():
            question = SurveyQuestionLikert.objects.get(key=key)
            answer_dict = {
                "assessment": self.assessment.pk,
                "question": question.pk,
                "choice": self.answers[key]["choice"],
                "explanation": self.answers[key]["explanation"],
            }
            try:
                answer = SurveyAnswerLikert.objects.get(
                    assessment=self.assessment, question=question
                )
                answer_serializer = SurveyAnswerLikertSerializer(
                    answer, data=answer_dict
                )
            except SurveyAnswerLikert.DoesNotExist:
                answer_serializer = SurveyAnswerLikertSerializer(data=answer_dict)

            answer_serializer.is_valid()
            if answer_serializer.errors:
                print(answer_serializer.errors)
                answer_serializer_errors.append(answer_serializer.errors)
            else:
                answer_serializers.append(answer_serializer)

        if answer_serializer_errors:
            # TODO: attach cell references to self.answers, and store with validations
            error = ingest_400(
                SURVEYANSWERLIKERTSERIALIZER,
                "invalid answers",
                {"errors": answer_serializer_errors},
            )
            self.validations.update(error)
            return

        with transaction.atomic():
            sid = transaction.savepoint()
            successful_save = False
            try:
                for a in answer_serializers:
                    a.save()
                successful_save = True
            finally:
                if dryrun is True or successful_save is False:
                    transaction.savepoint_rollback(sid)
                    if successful_save is False:
                        error = ingest_400(
                            ANSWER_SAVE, "error saving answers to database"
                        )
                        self.validations.update(error)
                else:
                    transaction.savepoint_commit(sid)
